# ============================================================
# src/utils/generate_excel.py — Trading Bot Trading Bot
#
# AMAÇ:
#   Botun yaptığı işlemlerin potansiyel kâr/zarar durumlarını,
#   sinyal gerekçelerini ve hatalardan ders çıkarmak için
#   post-mortem (hata analizi) değerlendirmelerini içeren
#   şık bir Excel raporu üretir.
#
# DEĞİŞİKLİK GEÇMİŞİ:
#   2026-06-04 | v1.0 | İlk Excel oluşturma aracı
#   2026-06-04 | v1.1 | Kapanış Tarihi kolonu eklendi, tarihler düzeltildi
# ============================================================

from __future__ import annotations

import os
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.utils.logger import get_logger
from src.utils.ai_analyzer import generate_post_mortem_analysis
from src.config.settings import get_settings

logger = get_logger(__name__)


def create_analysis_excel(output_path: str = "logs/islem_analiz_raporu.xlsx") -> None:
    """İşlem analizi ve hata değerlendirme Excel dosyasını oluşturur ve stillendirir."""
    
    # 1. logs/portfolio_state.json dosyasından canlı verileri oku
    import json
    state_path = "logs/portfolio_state.json"
    data = []
    
    if os.path.exists(state_path):
        try:
            with open(state_path, "r", encoding="utf-8") as f:
                state = json.load(f)
                
            open_positions = state.get("open_positions", {})
            orders = state.get("orders", [])
            
            # Açık pozisyonları ekle
            for sym, pos in open_positions.items():
                pnl_usdt = pos.get("pnl_usdt", 0.0)
                pnl_pct = pos.get("pnl_pct", 0.0)
                is_pending = pos.get("status", "active") == "pending"
                
                opened_at_str = pos.get("opened_at", "")
                try:
                    dt = datetime.fromisoformat(opened_at_str)
                    opened_at_formatted = dt.strftime("%Y-%m-%d %H:%M:%S")
                except Exception:
                    opened_at_formatted = opened_at_str
                
                if is_pending:
                    status_text = f"LİMİT ALIM BEKLENİYOR: Fiyat limit seviyesine gelince aktiflesecek. Limit Fiyat: ${pos['entry_price']:.4f}"
                    current_price_val = pos["entry_price"]
                else:
                    current_price_val = pos.get("current_price", pos["entry_price"])
                    if pnl_usdt >= 0:
                        status_text = f"AÇIK POZİSYON (KÂRDA): İşlem devam ediyor. Anlık PnL: ${pnl_usdt:+.2f} USDT ({pnl_pct*100:+.2f}%)"
                    else:
                        status_text = f"AÇIK POZİSYON (ZARARDA): İşlem devam ediyor. Anlık PnL: ${pnl_usdt:+.2f} USDT ({pnl_pct*100:+.2f}%)"
                
                data.append({
                    "Sembol": pos["symbol"],
                    "Yön": pos["side"].upper(),
                    "Giriş Tarihi": opened_at_formatted,
                    "Kapanış Tarihi": "Devam Ediyor" if not is_pending else "Limit Bekliyor",
                    "Giriş Fiyatı": pos["entry_price"],
                    "Stop-Loss (SL)": pos["stop_loss"],
                    "Take-Profit (TP)": pos["take_profit"],
                    "Çıkış Fiyatı (Örnek)": current_price_val,
                    "Gerekçe": f"Giriş: ${pos['entry_price']:.4f} | Anlık Fiyat: ${current_price_val:.4f}",
                    "Hata & Gelişim Analizi (Neden Kaybettik / Kazandık?)": status_text,
                    "is_open": True
                })
                
            # Kapatılmış pozisyonları ekle (exit orders ve entry orders olarak ayır)
            exit_orders = [o for o in orders if o.get("close_reason") is not None and o.get("status") == "filled"]
            entry_orders = [o for o in orders if o.get("close_reason") is None and o.get("status") == "filled"]
            
            for exit_ord in exit_orders:
                matching_entry = None
                expected_entry_side = "buy" if exit_ord["side"] == "sell" else "sell"
                for entry_ord in reversed(entry_orders):
                    if entry_ord["symbol"] == exit_ord["symbol"] and entry_ord["side"] == expected_entry_side and entry_ord["timestamp"] < exit_ord["timestamp"]:
                        matching_entry = entry_ord
                        break
                
                if matching_entry:
                    b_price = matching_entry["price"]  # entry price
                    s_price = exit_ord["price"]       # exit price
                    b_amount = matching_entry.get("amount", 0.0)
                    
                    # Gerçek PnL: eğer kaydedilmişse direkt al, yoksa hesapla
                    real_pnl_usdt = exit_ord.get("pnl_usdt")
                    real_pnl_pct = exit_ord.get("pnl_pct")
                    
                    if real_pnl_usdt is None:
                        if expected_entry_side == "buy":
                            real_pnl_usdt = (s_price - b_price) * b_amount
                        else:
                            real_pnl_usdt = (b_price - s_price) * b_amount
                            
                    if real_pnl_pct is None:
                        if expected_entry_side == "buy":
                            real_pnl_pct = (s_price - b_price) / b_price if b_price > 0 else 0.0
                        else:
                            real_pnl_pct = (b_price - s_price) / b_price if b_price > 0 else 0.0
                    
                    try:
                        dt_exit = datetime.fromisoformat(exit_ord["timestamp"])
                        exit_time_formatted = dt_exit.strftime("%Y-%m-%d %H:%M:%S")
                    except Exception:
                        exit_time_formatted = exit_ord["timestamp"]

                    try:
                        dt_entry = datetime.fromisoformat(matching_entry["timestamp"])
                        entry_time_formatted = dt_entry.strftime("%Y-%m-%d %H:%M:%S")
                    except Exception:
                        entry_time_formatted = matching_entry["timestamp"]
 
                    close_reason = exit_ord.get("close_reason", "KAPANDI")
                    direction_str = "LONG" if expected_entry_side == "buy" else "SHORT"
                    gerekce = f"Kapanış: {close_reason} | Giriş: ${b_price:.4f} | Çıkış: ${s_price:.4f}"
                        
                    data.append({
                        "Sembol": exit_ord["symbol"],
                        "Yön": direction_str,
                        "Giriş Tarihi": entry_time_formatted,
                        "Kapanış Tarihi": exit_time_formatted,
                        "Giriş Fiyatı": b_price,
                        "Stop-Loss (SL)": matching_entry.get("stop_loss", 0.0) or (b_price * 0.9 if direction_str == "LONG" else b_price * 1.1),
                        "Take-Profit (TP)": matching_entry.get("take_profit", 0.0) or (b_price * 1.1 if direction_str == "LONG" else b_price * 0.9),
                        "Çıkış Fiyatı (Örnek)": s_price,
                        "Gerekçe": gerekce,
                        "pnl_usdt_real": real_pnl_usdt,
                        "pnl_pct_real": real_pnl_pct,
                        "indicator_summary": exit_ord.get("indicator_summary"),
                        "is_open": False
                    })
        except Exception as e:
            logger.error(f"portfolio_state.json okunurken hata: {e}")
            
    # Eğer portföy boşsa veya dosya yoksa, yedek statik verileri doldur (demo amaçlı)
    if not data:
        data = [
            {
                "Sembol": "OPN/USDT",
                "Yön": "LONG",
                "Giriş Tarihi": "2026-06-03 21:35:00",
                "Kapanış Tarihi": "2026-06-04 09:52:00",
                "Giriş Fiyatı": 0.26,
                "Stop-Loss (SL)": 0.22,
                "Take-Profit (TP)": 0.32,
                "Çıkış Fiyatı (Örnek)": 0.22,
                "Gerekçe": "Bollinger Squeeze daralması sonrası yukarı kırılım beklentisi.",
                "indicator_summary": {
                    "Bollinger_Bands": "Squeeze (0.015)",
                    "Trend_1d": "Bearish",
                    "Trend_4h": "Fake Bullish Breakout",
                    "RSI": "48"
                },
                "is_open": False
            },
            {
                "Sembol": "STO/USDT",
                "Yön": "LONG",
                "Giriş Tarihi": "2026-06-03 21:35:00",
                "Kapanış Tarihi": "2026-06-04 09:52:00",
                "Giriş Fiyatı": 0.07,
                "Stop-Loss (SL)": 0.06,
                "Take-Profit (TP)": 0.09,
                "Çıkış Fiyatı (Örnek)": 0.09,
                "Gerekçe": "RSI Uyuşmazlığı (Divergence) + MACD Alış Kesişimi.",
                "indicator_summary": {
                    "RSI": "32 (Divergence)",
                    "MACD": "Bullish Cross",
                    "Volume_Ratio": "1.4x"
                },
                "is_open": False
            }
        ]

    from openpyxl import load_workbook
    
    today_str = datetime.now().strftime("%Y-%m-%d")
    
    if os.path.exists(output_path):
        try:
            wb = load_workbook(output_path)
            logger.info(f"Mevcut Excel dosyası yüklendi: {output_path}")
        except Exception as e:
            logger.error(f"Excel dosyası yüklenirken hata: {e}. Yeni dosya oluşturuluyor.")
            wb = Workbook()
    else:
        wb = Workbook()
 
    if today_str in wb.sheetnames:
        del wb[today_str]
        
    ws = wb.create_sheet(title=today_str)
    
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]
        
    ws.views.sheetView[0].showGridLines = True

    # 2. Başlık ve Tablo Hazırlığı
    headers = [
        "Sembol", "Yön", "Giriş Tarihi", "Kapanış Tarihi", "Giriş Fiyatı ($)", 
        "Stop-Loss ($)", "Take-Profit ($)", "Çıkış Fiyatı ($)", 
        "Potansiyel PnL ($)", "Potansiyel PnL (%)", 
        "Sinyal Gerekçesi", "Hata & Gelişim Analizi (Post-Mortem)"
    ]

    title_font = Font(name="Calibri", size=16, bold=True, color="1F497D")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    
    thin_border_side = Side(border_style="thin", color="D3D3D3")
    thin_border = Border(
        left=thin_border_side, right=thin_border_side, 
        top=thin_border_side, bottom=thin_border_side
    )
    double_bottom_border = Border(
        top=thin_border_side, 
        bottom=Side(border_style="double", color="000000")
    )

    ws["A1"] = "Trading Bot TRADING BOT — İŞLEM POST-MORTEM & HATA ANALİZ RAPORU"
    ws["A1"].font = title_font
    ws.row_dimensions[1].height = 30
    
    ws["A2"] = f"Rapor Oluşturulma Tarihi: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws.row_dimensions[2].height = 18

    ws.row_dimensions[4].height = 25
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 3. Verileri Doldur ve Hesaplamaları Excel Formülü Olarak Yaz
    start_row = 5
    for i, row_data in enumerate(data):
        r = start_row + i
        ws.row_dimensions[r].height = 65

        ws.cell(row=r, column=1, value=row_data["Sembol"]).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=r, column=2, value=row_data["Yön"]).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=r, column=3, value=row_data["Giriş Tarihi"]).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=r, column=4, value=row_data["Kapanış Tarihi"]).alignment = Alignment(horizontal="center", vertical="center")
        
        ws.cell(row=r, column=5, value=row_data["Giriş Fiyatı"]).number_format = "$#,##0.0000"
        ws.cell(row=r, column=6, value=row_data["Stop-Loss (SL)"]).number_format = "$#,##0.0000"
        ws.cell(row=r, column=7, value=row_data["Take-Profit (TP)"]).number_format = "$#,##0.0000"
        ws.cell(row=r, column=8, value=row_data["Çıkış Fiyatı (Örnek)"]).number_format = "$#,##0.0000"
        
        ws.cell(row=r, column=9, value=f"=1000 * (H{r} - E{r}) / E{r}").number_format = "$#,##0.00"
        ws.cell(row=r, column=10, value=f"=(H{r} - E{r}) / E{r}").number_format = "0.00%"

        ws.cell(row=r, column=11, value=row_data["Gerekçe"]).alignment = Alignment(wrap_text=True, vertical="center")
        
        # Eğer açık pozisyonsa, yapay zeka çağırmak yerine doğrudan canlı PnL durumunu yazdırıyoruz
        if row_data.get("is_open", False):
            ai_analiz = row_data["Hata & Gelişim Analizi (Neden Kaybettik / Kazandık?)"]
        else:
            # Kapatılmış pozisyonsa, yapay zeka post-mortem analizini çağırıyoruz
            pnl_pct = (row_data["Çıkış Fiyatı (Örnek)"] - row_data["Giriş Fiyatı"]) / row_data["Giriş Fiyatı"]
            pnl_usdt = 1000.0 * pnl_pct
            
            ai_analiz = generate_post_mortem_analysis(
                symbol=row_data["Sembol"],
                side=row_data["Yön"],
                entry_price=row_data["Giriş Fiyatı"],
                exit_price=row_data["Çıkış Fiyatı (Örnek)"],
                stop_loss=row_data["Stop-Loss (SL)"],
                take_profit=row_data["Take-Profit (TP)"],
                pnl_usdt=pnl_usdt,
                pnl_pct=pnl_pct,
                reason=row_data["Gerekçe"],
                indicator_summary=row_data.get("indicator_summary")
            )
            ai_analiz = f"KAPANDI: {ai_analiz}"
        
        ws.cell(row=r, column=12, value=ai_analiz).alignment = Alignment(wrap_text=True, vertical="center")

        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            cell.font = Font(name="Calibri", size=10)
            if c in [5, 6, 7, 8, 9, 10]:
                cell.alignment = Alignment(horizontal="right", vertical="center")

    # 4. Toplam (Summary) Satırı Ekle
    total_row = start_row + len(data)
    ws.row_dimensions[total_row].height = 25
    
    ws.cell(row=total_row, column=1, value="TOPLAM").font = Font(name="Calibri", size=10, bold=True)
    ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
    
    total_pnl_cell = ws.cell(row=total_row, column=9, value=f"=SUM(I5:I{total_row-1})")
    total_pnl_cell.number_format = "$#,##0.00"
    total_pnl_cell.font = Font(name="Calibri", size=10, bold=True)
    total_pnl_cell.alignment = Alignment(horizontal="right", vertical="center")

    avg_pnl_cell = ws.cell(row=total_row, column=10, value=f"=AVERAGE(J5:J{total_row-1})")
    avg_pnl_cell.number_format = "0.00%"
    avg_pnl_cell.font = Font(name="Calibri", size=10, bold=True)
    avg_pnl_cell.alignment = Alignment(horizontal="right", vertical="center")

    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=total_row, column=c)
        cell.border = double_bottom_border

    # 5. Kâr/Zarar Hücrelerini Koşullu Renklendirme (Pozitif yeşil, negatif kırmızı)
    green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    red_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    green_font = Font(name="Calibri", size=10, color="375623")
    red_font = Font(name="Calibri", size=10, color="C65911")

    for i in range(len(data)):
        r = start_row + i
        entry_val = data[i]["Giriş Fiyatı"]
        exit_val = data[i]["Çıkış Fiyatı (Örnek)"]
        
        pnl_val_cell = ws.cell(row=r, column=9)
        pnl_pct_cell = ws.cell(row=r, column=10)
        
        if exit_val >= entry_val:
            pnl_val_cell.fill = green_fill
            pnl_val_cell.font = green_font
            pnl_pct_cell.fill = green_fill
            pnl_pct_cell.font = green_font
        else:
            pnl_val_cell.fill = red_fill
            pnl_val_cell.font = red_font
            pnl_pct_cell.fill = red_fill
            pnl_pct_cell.font = red_font

    # 6. Sütun Genişliklerini Ayarla (Yazıların sığması için otomatik genişlik)
    column_widths = {
        "A": 12,  # Sembol
        "B": 8,   # Yön
        "C": 20,  # Giriş Tarihi
        "D": 20,  # Kapanış Tarihi
        "E": 15,  # Giriş Fiyatı
        "F": 15,  # SL
        "G": 15,  # TP
        "H": 15,  # Çıkış Fiyatı
        "I": 18,  # PnL $
        "J": 15,  # PnL %
        "K": 35,  # Gerekçe
        "L": 55,  # Hata Analizi
    }

    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Kaydet
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    try:
        wb.save(output_path)
        logger.info(f"[KAYDET] Post-Mortem Excel Raporu olusturuldu: {output_path}")
    except PermissionError:
        timestamp = datetime.now().strftime("%H%M%S")
        backup_path = output_path.replace(".xlsx", f"_yedek_{timestamp}.xlsx")
        logger.warning(
            f"[UYARI] Hata: '{output_path}' dosyasi acik oldugu icin yazilamadi (Kilitli). "
            f"Rapor yedek olarak suraya kaydedildi: {backup_path}"
        )
        wb.save(backup_path)


if __name__ == "__main__":
    create_analysis_excel()
